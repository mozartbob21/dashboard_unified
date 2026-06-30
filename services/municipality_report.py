
from typing import List
from urllib.parse import quote
from pathlib import Path
import csv
import io
import json
import re
from functools import lru_cache
from datetime import datetime
from html import escape

from fastapi import APIRouter, Query, Request
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates

EXCLUDED_MUNICIPALITY_REPORT_FILE_NAMES = {
    "dashboard_state.json",
    "final_result.json",
    "summary.json",
    "cache.json",
    "state.json",
}


EXCLUDED_MUNICIPALITY_REPORT_PATH_PARTS = {
    "debug",
    "responses",
    "state",
    "__pycache__",
    ".git",
    ".venv",
    "venv",
    "node_modules",
}


MUNICIPALITY_FIELD_HINTS = [
    "municipality",
    "муниципалитет",
    "муниципальное образование",
    "городской округ",
    "г.о.",
    "го",
    "territory",
    "территория",
    "okrug",
    "district",
    "район",
]


ADDRESS_FIELD_HINTS = [
    "address",
    "адрес",
    "street",
    "улица",
    "ул",
    "location",
    "местоположение",
]



try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
except Exception:
    colors = None


router = APIRouter()

BASE_DIR = Path(__file__).resolve().parents[1]
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

SCAN_DIRS = [
    BASE_DIR / "data",
    BASE_DIR / "generated",
]

ALLOWED_EXTENSIONS = {
    ".json",
    ".csv",
    ".xlsx",
}


@lru_cache(maxsize=200_000)
def _normalize_text_cached(value):
    value = str(value or "")
    value = value.strip().lower()
    value = value.replace("ё", "е")
    value = re.sub(r"[^а-яa-z0-9]+", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_text(value):
    try:
        return _normalize_text_cached(value)
    except TypeError:
        return _normalize_text_cached(str(value))


def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value)


def flatten_json_records(obj):
    records = []

    if isinstance(obj, list):
        for item in obj:
            records.extend(flatten_json_records(item))
        return records

    if isinstance(obj, dict):
        # Если словарь похож на одну запись
        primitive_count = 0
        nested_values = []

        for value in obj.values():
            if isinstance(value, (dict, list)):
                nested_values.append(value)
            else:
                primitive_count += 1

        if primitive_count > 0:
            records.append(obj)

        for value in nested_values:
            records.extend(flatten_json_records(value))

    return records


def row_matches_municipality(row, municipality_name):
    target = normalize_text(municipality_name)
    if not target:
        return False

    if isinstance(row, dict):
        haystack = " ".join(safe_str(v) for v in row.values())
    else:
        haystack = safe_str(row)

    haystack_norm = normalize_text(haystack)
    return target in haystack_norm


def read_json_file(path):
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="cp1251", errors="ignore")

    data = json.loads(text)
    records = flatten_json_records(data)
    return records


def read_csv_file(path):
    rows = []

    encodings = ["utf-8-sig", "utf-8", "cp1251"]
    last_error = None

    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(dict(row))
            return rows
        except Exception as e:
            last_error = e
            rows = []

    raise last_error


def read_xlsx_file(path):
    if load_workbook is None:
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []

    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue

        headers = [safe_str(x).strip() or f"Колонка {i+1}" for i, x in enumerate(data[0])]

        for raw_row in data[1:]:
            item = {}
            for i, value in enumerate(raw_row):
                key = headers[i] if i < len(headers) else f"Колонка {i+1}"
                item[key] = value
            item["_sheet"] = ws.title
            rows.append(item)

    return rows


def read_records_from_file(path):
    suffix = path.suffix.lower()

    if suffix == ".json":
        return read_json_file(path)

    if suffix == ".csv":
        return read_csv_file(path)

    if suffix == ".xlsx":
        return read_xlsx_file(path)

    return []


MODULE_LABELS = {
    "utnkr": "УТН и капитальный ремонт",
    "watercontrol": "Водный контроль",
    "auth": "Авторизация",
    "cameras": "Камеры видеонаблюдения",
    "cds": "ЦДС (диспетчерская служба)",
    "edo": "Электронный документооборот",
    "overdue": "Просроченные задачи",
    "planfix": "Planfix",
    "data": "Данные",
}


def humanize_module(module_name):
    s = str(module_name).strip()
    if s in MODULE_LABELS:
        return MODULE_LABELS[s]
    pretty = s.replace("_", " ").replace("-", " ").strip()
    return pretty[:1].upper() + pretty[1:] if pretty else s


def detect_module_name(path):
    try:
        rel = path.relative_to(BASE_DIR)
    except Exception:
        rel = path

    parts = rel.parts

    if len(parts) >= 2:
        if parts[0] in {"data", "generated", "static"}:
            if len(parts) >= 3:
                return parts[1]
            return path.stem

    return path.stem


def collect_municipality_data(municipality_name):
    groups = {}
    errors = []

    for scan_dir in SCAN_DIRS:
        if not scan_dir.exists():
            continue

        for path in scan_dir.rglob("*"):
            if not path.is_file():
                continue

            if path.suffix.lower() not in ALLOWED_EXTENSIONS:
                continue

            # Пропускаем временные файлы Excel
            if path.name.startswith("~$"):
                continue

            # Пропускаем исходный .xlsx, если рядом есть обработанный result.json
            # (это одни и те же данные: xlsx — выгрузка из 1С, json — её результат)
            if path.suffix.lower() == ".xlsx" and (path.parent / "result.json").exists():
                continue

            try:
                records = read_records_from_file(path)
            except Exception as e:
                errors.append({
                    "file": str(path.relative_to(BASE_DIR)),
                    "error": str(e),
                })
                continue

            matched = []
            for row in records:
                if row_matches_municipality(row, municipality_name):
                    matched.append(row)

            if matched:
                module_name = detect_module_name(path)
                base_key = humanize_module(module_name)
                key = base_key
                _suffix = 2
                while key in groups:
                    key = f"{base_key} ({_suffix})"
                    _suffix += 1

                groups[key] = {
                    "module": module_name,
                    "file": str(path.relative_to(BASE_DIR)),
                    "count": len(matched),
                    "rows": matched,
                }

    total = sum(group["count"] for group in groups.values())

    return {
        "municipality": municipality_name,
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "groups": groups,
        "errors": errors,
        "total": total,
    }


FIELD_LABELS = {
    "id": "ID",
    "name": "Наименование",
    "title": "Название",
    "address": "Адрес",
    "status": "Статус",
    "responsible_name": "Ответственный",
    "responsible": "Ответственный",
    "is_edited": "Изменено",
    "edited": "Изменено",
    "created_at": "Создано",
    "updated_at": "Обновлено",
    "date": "Дата",
    "deadline": "Срок",
    "comment": "Комментарий",
    "description": "Описание",
    "category": "Категория",
    "type": "Тип",
    "area": "Площадь",
    "municipality": "Муниципалитет",
    "district": "Район",
    "result": "Результат",
    "value": "Значение",
    "count": "Количество",
    "phone": "Телефон",
    "email": "E-mail",
    "url": "Ссылка",
    "link": "Ссылка",
    "priority": "Приоритет",
    "risk": "Риск",
    "score": "Оценка",
    "note": "Примечание",
    "annotation": "Аннотация",
    "action": "Действие",
    "task_id": "ID задачи",
    "grbs": "ГРБС",
    "situation": "Ситуация на объекте",
    "raw_row_text": "Исходный текст",
    "row_number": "№ строки",
    "column": "Колонка",
    "column_index": "Индекс колонки",
    "applicant": "Заявитель",
    "organization": "Организация",
    "metric_name": "Наименование показателя",
    "metric": "Показатель",
    "current_value": "Текущее значение",
    "target_value": "Целевое значение",
    "delay_days": "Дней просрочки",
    "responsible_phone": "Телефон ответственного",
    "responsible_position": "Должность ответственного",
    "period": "Период",
    "unit": "Ед. изм.",
    "fact": "Факт",
    "plan": "План",
    "percent": "Процент",
}

VALUE_LABELS = {
    "True": "Да",
    "False": "Нет",
    "true": "Да",
    "false": "Нет",
    "None": "—",
    "null": "—",
    "": "—",
    "nan": "—",
    "risk": "Риск",
    "ok": "Норма",
    "done": "Выполнено",
    "in_progress": "В работе",
    "pending": "Ожидает",
}


def humanize_field(key):
    s = str(key).strip()
    if s in FIELD_LABELS:
        return FIELD_LABELS[s]
    norm = s.lower().replace(" ", "_").strip()
    if norm in FIELD_LABELS:
        return FIELD_LABELS[norm]
    pretty = s.replace("_", " ").strip()
    return pretty[:1].upper() + pretty[1:] if pretty else s


def humanize_value(value):
    s = str(value).strip()
    return VALUE_LABELS.get(s, s)


HIDDEN_COLUMNS = {
    "raw_row_text",
    "column_index",
    "_sheet",
}


def get_all_columns(rows, max_columns=12):
    columns = []

    for row in rows:
        if isinstance(row, dict):
            for key in row.keys():
                if key in HIDDEN_COLUMNS:
                    continue
                if key not in columns:
                    columns.append(key)

        if len(columns) >= max_columns:
            break

    return columns[:max_columns]


def register_pdf_font():
    font_candidates = [
        BASE_DIR / "static" / "fonts" / "DejaVuSans.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
    ]

    for font_path in font_candidates:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("ReportFont", str(font_path)))
                return "ReportFont"
            except Exception:
                pass

    return "Helvetica"


def build_pdf_bytes(report):
    if colors is None:
        raise RuntimeError("Библиотека reportlab не установлена. Выполни: python3 -m pip install reportlab")

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Отчёт по муниципалитету {report['municipality']}",
    )

    font_name = register_pdf_font()

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "MunicipalityTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        leading=22,
        alignment=1,
        spaceAfter=10,
    )

    subtitle_style = ParagraphStyle(
        "MunicipalitySubtitle",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=13,
        spaceAfter=8,
    )

    h2_style = ParagraphStyle(
        "MunicipalityH2",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=13,
        leading=16,
        spaceBefore=8,
        spaceAfter=6,
    )

    cell_style = ParagraphStyle(
        "MunicipalityCell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7,
        leading=9,
    )

    story = []

    story.append(Paragraph(f"Сводный отчёт по муниципалитету: {escape(report['municipality'])}", title_style))
    story.append(Paragraph(f"Дата формирования: {escape(report['generated_at'])}", subtitle_style))
    story.append(Paragraph(f"Всего найдено записей: {report['total']}", subtitle_style))
    story.append(Spacer(1, 4 * mm))

    if not report["groups"]:
        story.append(Paragraph("Данные по указанному муниципалитету не найдены.", subtitle_style))
    else:
        for index, (group_name, group) in enumerate(report["groups"].items(), start=1):
            if index > 1:
                story.append(PageBreak())

            story.append(Paragraph(f"{index}. {escape(humanize_module(group.get('module', group_name)))}", h2_style))
            story.append(Paragraph(f"Найдено записей: {group['count']}", subtitle_style))

            rows = group["rows"]
            preview_rows = rows[:40]
            columns = get_all_columns(preview_rows, max_columns=8)

            if not columns:
                story.append(Paragraph("Нет колонок для отображения.", subtitle_style))
                continue

            table_data = []
            table_data.append([Paragraph(escape(humanize_field(col)), cell_style) for col in columns])

            for row in preview_rows:
                line = []
                for col in columns:
                    value = safe_str(row.get(col, "")) if isinstance(row, dict) else safe_str(row)
                    value = humanize_value(value)
                    value = value[:500]
                    line.append(Paragraph(escape(value), cell_style))
                table_data.append(line)

            col_width = (landscape(A4)[0] - 24 * mm) / max(len(columns), 1)
            table = Table(table_data, colWidths=[col_width] * len(columns), repeatRows=1)

            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))

            story.append(table)

            if group["count"] > len(preview_rows):
                story.append(Spacer(1, 3 * mm))
                story.append(Paragraph(
                    f"В PDF показаны первые {len(preview_rows)} записей из {group['count']}. Полный просмотр доступен на HTML-странице.",
                    subtitle_style
                ))

    if report["errors"]:
        story.append(PageBreak())
        story.append(Paragraph("Ошибки чтения файлов", h2_style))

        error_data = [[
            Paragraph("Файл", cell_style),
            Paragraph("Ошибка", cell_style),
        ]]

        for error in report["errors"][:100]:
            error_data.append([
                Paragraph(escape(error["file"]), cell_style),
                Paragraph(escape(error["error"][:800]), cell_style),
            ])

        table = Table(error_data, colWidths=[100 * mm, 160 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FEE2E2")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)

    doc.build(story)

    buffer.seek(0)
    return buffer.getvalue()




# ===============================
# COMPACT HTML VIEW HELPERS
# ===============================

HIDDEN_HTML_COLUMNS = {
    "rows",
    "personal_messages",
    "public_chat_message",
    "screenshot_paths",
    "screenshots",
    "raw",
    "html",
    "payload",
    "data",
    "items",
    "children",
    "nested",
    "debug",
    "trace",
}

PREFERRED_HTML_COLUMNS = [
    "municipality",
    "organization",
    "name",
    "title",
    "status",
    "address",
    "created_at",
    "date",
    "deadline",
    "overdue",
    "summary",
    "summary_message",
    "message",
    "responsible_name",
    "responsible_phone",
    "source_url",
    "redmine_url",
    "screenshot_path",
]


def compact_value(value, limit=180):
    text = safe_str(value)
    text = re.sub(r"\s+", " ", text).strip()

    if len(text) > limit:
        return text[:limit].rstrip() + "…"

    return text


def is_bad_html_column(column):
    col = str(column or "").strip().lower()

    if col in HIDDEN_HTML_COLUMNS:
        return True

    if col.startswith("_"):
        return True

    if "screenshot_paths" in col:
        return True

    if "personal_messages" in col:
        return True

    if col in {"rows", "row", "raw_json"}:
        return True

    return False


def prepare_group_for_html(group, max_rows=10, max_columns=7):
    rows = group.get("rows", []) or []

    all_columns = []
    for row in rows[:50]:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in all_columns and not is_bad_html_column(key):
                    all_columns.append(key)

    preferred = [c for c in PREFERRED_HTML_COLUMNS if c in all_columns]
    others = [c for c in all_columns if c not in preferred]

    columns = (preferred + others)[:max_columns]

    compact_rows = []
    for row in rows[:max_rows]:
        item = {}
        for col in columns:
            if isinstance(row, dict):
                item[col] = compact_value(row.get(col, ""))
            else:
                item[col] = compact_value(row)
        compact_rows.append(item)

    result = dict(group)
    result["columns"] = columns
    result["compact_rows"] = compact_rows
    result["shown_count"] = len(compact_rows)
    result["hidden_count"] = max(0, int(group.get("count", 0)) - len(compact_rows))

    return result


def prepare_report_for_html(report):
    if not report:
        return report

    prepared = dict(report)
    prepared_groups = {}

    for group_name, group in report.get("groups", {}).items():
        prepared_groups[group_name] = prepare_group_for_html(group)

    prepared["groups"] = prepared_groups
    return prepared


@router.get("/municipality-report", response_class=HTMLResponse)
async def municipality_report_page(
    request: Request,
    name: str = Query("", description="Название муниципалитета"),
    blocks: List[str] = Query(default=[], description="Фильтр по блокам"),
):
    report = None
    selected_blocks = [b for b in blocks if b in VALID_BLOCK_KEYS]

    if name.strip():
        report = collect_municipality_data(name.strip())
        report = prepare_report_for_html(report, selected_blocks=selected_blocks)

    municipalities = list_available_municipalities()

    return templates.TemplateResponse(
        request=request,
        name="municipality_report.html",
        context={
            "request": request,
            "name": name.strip(),
            "report": report,
            "municipalities": municipalities,
            "all_blocks": REPORT_BLOCKS,
            "selected_blocks": selected_blocks,
        },
    )


@router.get("/municipality-report.pdf")
async def municipality_report_pdf(
    name: str = Query(..., description="Название муниципалитета"),
):
    municipality = name.strip()

    if not municipality:
        return Response("Не указан муниципалитет", status_code=400)

    report = collect_municipality_data(municipality)
    pdf_bytes = build_pdf_bytes(report)

    safe_name = re.sub(r"[^а-яА-Яa-zA-Z0-9_-]+", "_", municipality).strip("_")
    filename = f"{safe_name}_svodnyy_otchet.pdf"

    # ASCII-имя для старых клиентов (кириллицу транслитерируем грубо в "_")
    ascii_filename = re.sub(r"[^a-zA-Z0-9_.-]+", "_", filename).strip("_") or "report.pdf"
    # RFC 5987: percent-encoded UTF-8 имя
    utf8_filename = quote(filename)
    headers = {
        "Content-Disposition": (
            f"attachment; filename=\"{ascii_filename}\"; "
            f"filename*=UTF-8''{utf8_filename}"
        )
    }

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers=headers,
    )


# ============================================================
# FACT-BASED MUNICIPALITY REPORT VIEW
# ============================================================

TECHNICAL_FACT_COLUMNS = {
    "rows",
    "row",
    "raw",
    "raw_json",
    "json",
    "html",
    "payload",
    "data",
    "items",
    "children",
    "nested",
    "debug",
    "trace",
    "screenshots",
    "screenshot_paths",
    "personal_messages",
    "public_chat_message",
    "_sheet",
}

IMPORTANT_FACT_KEYS = [
    "municipality",
    "territory",
    "organization",
    "org",
    "name",
    "title",
    "object",
    "address",
    "status",
    "deadline",
    "date",
    "created_at",
    "responsible",
    "responsible_name",
    "phone",
    "comment",
    "problem",
    "issue",
    "reason",
    "url",
    "source_url",
    "redmine_url",
]

EMPTY_MARKERS = {
    "",
    "-",
    "—",
    "none",
    "null",
    "nan",
    "нет",
    "не указано",
    "не указан",
    "не заполнено",
    "не заполнен",
    "отсутствует",
}


def fact_norm(value):
    return normalize_text(value)


def fact_str(value, limit=220):
    text = safe_str(value)
    text = re.sub(r"\s+", " ", text).strip()

    if len(text) > limit:
        return text[:limit].rstrip() + "…"

    return text


def is_empty_fact_value(value):
    text = fact_norm(value)
    return text in EMPTY_MARKERS


def is_technical_fact_key(key):
    key_norm = str(key or "").strip().lower()

    if not key_norm:
        return True

    if key_norm in TECHNICAL_FACT_COLUMNS:
        return True

    if key_norm.startswith("_"):
        return True

    bad_parts = [
        "screenshot_paths",
        "personal_messages",
        "raw_json",
        "payload",
        "debug",
        "trace",
    ]

    return any(part in key_norm for part in bad_parts)


def row_text_for_facts(row):
    if isinstance(row, dict):
        return " ".join(fact_str(v, 500) for k, v in row.items() if not is_technical_fact_key(k))

    return fact_str(row, 1000)


def classify_fact_module(group_name, file_path):
    text = fact_norm(f"{group_name} {file_path}")

    if any(x in text for x in ["1с", "1c", "obrasheniya_1c", "appeals_1c"]):
        return {"key": "appeals_1c", "title": "Обращения 1С",
                "subtitle": "Обращения из системы 1С", "icon": "🧾", "tone": "info"}

    if any(x in text for x in ["эмпат", "empath", "empathy", "тон ответа", "вежлив"]):
        return {"key": "empathy", "title": "Эмпатичные ответы",
                "subtitle": "Качество и тональность ответов", "icon": "💬", "tone": "info"}

    if any(x in text for x in ["overdue", "просроч", "deadline", "срок"]):
        return {"key": "overdue", "title": "Просроченные задачи",
                "subtitle": "Контроль исполнения и нарушенных сроков", "icon": "⏰", "tone": "danger"}

    if any(x in text for x in ["edo", "эдо", "заполн", "completeness"]):
        return {"key": "edo", "title": "Заполненность данных",
                "subtitle": "Контроль полноты и корректности данных", "icon": "▦", "tone": "warning"}

    if any(x in text for x in ["water", "вод", "watercontrol", "контроль воды"]):
        return {"key": "water", "title": "Контроль воды",
                "subtitle": "Заполненность карточек и обязательных полей", "icon": "💧", "tone": "info"}

    if any(x in text for x in ["camera", "камер", "video", "ffmpeg", "stream"]):
        return {"key": "cameras", "title": "Проверка камер",
                "subtitle": "Работоспособность камер и видеопотоков", "icon": "📹", "tone": "danger"}

    if any(x in text for x in ["utnkr", "утнкр", "технадзор", "tehnadzor"]):
        return {"key": "utnkr", "title": "Технадзор УТНКР",
                "subtitle": "Просрочки, статусы и проблематика объектов", "icon": "🏗", "tone": "warning"}

    if any(x in text for x in ["appeal", "обращ", "ответ", "жалоб", "cds", "цдс", "диспетчер"]):
        return {"key": "appeals", "title": "Обращения",
                "subtitle": "Анализ обращений и подготовка ответов", "icon": "✉️", "tone": "info"}

    return {"key": "other", "title": "Прочие данные",
            "subtitle": "Дополнительные найденные сведения", "icon": "📄", "tone": "neutral"}


REPORT_BLOCKS = [
    {"key": "edo",        "title": "Заполненность данных", "icon": "▦"},
    {"key": "overdue",    "title": "Просроченные задачи",  "icon": "⏰"},
    {"key": "water",      "title": "Контроль воды",        "icon": "💧"},
    {"key": "utnkr",      "title": "Технадзор УТНКР",      "icon": "🏗"},
    {"key": "cameras",    "title": "Проверка камер",       "icon": "📹"},
    {"key": "empathy",    "title": "Эмпатичные ответы",    "icon": "💬"},
    {"key": "appeals_1c", "title": "Обращения 1С",         "icon": "🧾"},
    {"key": "appeals",    "title": "Обращения",            "icon": "✉️"},
    {"key": "other",      "title": "Прочие данные",        "icon": "📄"},
]

VALID_BLOCK_KEYS = {b["key"] for b in REPORT_BLOCKS}


def detect_overdue_row(row):
    text = fact_norm(row_text_for_facts(row))

    overdue_words = [
        "просроч",
        "overdue",
        "нарушен срок",
        "срок нарушен",
        "задерж",
    ]

    if any(word in text for word in overdue_words):
        return True

    if isinstance(row, dict):
        for key, value in row.items():
            key_norm = fact_norm(key)
            value_norm = fact_norm(value)

            if any(x in key_norm for x in ["overdue", "просроч"]):
                if value_norm not in EMPTY_MARKERS and value_norm not in {"0", "false", "нет", "no"}:
                    return True

            if any(x in key_norm for x in ["days_overdue", "delay", "дней просрочки", "просрочка дней"]):
                try:
                    number = float(str(value).replace(",", "."))
                    if number > 0:
                        return True
                except Exception:
                    pass

    return False


def detect_camera_problem_row(row):
    text = fact_norm(row_text_for_facts(row))

    problem_words = [
        "нет видеопотока",
        "видеопоток отсутствует",
        "нет потока",
        "не подключ",
        "камера отсутствует",
        "offline",
        "off line",
        "недоступ",
        "ошибка подключения",
        "connection error",
        "timeout",
        "не работает",
    ]

    return any(word in text for word in problem_words)


def detect_risk_row(row):
    text = fact_norm(row_text_for_facts(row))

    risk_words = [
        "красный",
        "red",
        "высокий риск",
        "critical",
        "критич",
        "авар",
        "проблем",
        "наруш",
    ]

    return any(word in text for word in risk_words)


def detect_empty_fields(row):
    if not isinstance(row, dict):
        return []

    result = []

    for key, value in row.items():
        if is_technical_fact_key(key):
            continue

        if is_empty_fact_value(value):
            result.append(str(key))

    return result


def pick_first_value(row, candidates):
    if not isinstance(row, dict):
        return ""

    normalized_map = {fact_norm(k): k for k in row.keys()}

    for candidate in candidates:
        candidate_norm = fact_norm(candidate)

        for norm_key, original_key in normalized_map.items():
            if candidate_norm == norm_key or candidate_norm in norm_key:
                value = row.get(original_key)
                if not is_empty_fact_value(value):
                    return fact_str(value, 180)

    return ""


def object_title_from_row(row):
    title = pick_first_value(row, [
        "organization",
        "организация",
        "org",
        "name",
        "title",
        "object",
        "объект",
        "address",
        "адрес",
        "муниципалитет",
        "municipality",
    ])

    if title:
        return title

    if isinstance(row, dict):
        for key, value in row.items():
            if not is_technical_fact_key(key) and not is_empty_fact_value(value):
                return fact_str(value, 180)

    return "Запись без названия"


def row_badges(row):
    badges = []

    status = pick_first_value(row, ["status", "статус"])
    if status:
        badges.append(status)

    deadline = pick_first_value(row, ["deadline", "срок", "дата"])
    if deadline:
        badges.append(f"Срок: {deadline}")

    responsible = pick_first_value(row, ["responsible", "ответственный", "исполнитель"])
    if responsible:
        badges.append(f"Ответственный: {responsible}")

    address = pick_first_value(row, ["address", "адрес"])
    if address:
        badges.append(address)

    return badges[:4]


def row_problem_text(row):
    if detect_overdue_row(row):
        return "Есть признак просрочки или нарушения срока."

    if detect_camera_problem_row(row):
        return "Есть проблема с камерой или видеопотоком."

    missing = detect_empty_fields(row)
    if missing:
        return "Не заполнены поля: " + ", ".join(missing[:5])

    if detect_risk_row(row):
        return "Есть признаки риска, нарушения или проблемного статуса."

    problem = pick_first_value(row, ["problem", "issue", "reason", "comment", "проблема", "причина", "комментарий"])
    if problem:
        return problem

    return "Запись найдена по выбранному муниципалитету."


def add_fact_metric(metrics, label, value, tone="neutral", hint=""):
    metrics.append({
        "label": label,
        "value": value,
        "tone": tone,
        "hint": hint,
    })


def build_fact_card(module_info, group_name, group):
    rows = group.get("rows", []) or []
    total = len(rows)

    overdue_rows = []
    camera_problem_rows = []
    risk_rows = []
    rows_with_missing = []

    missing_counter = {}

    for row in rows:
        if detect_overdue_row(row):
            overdue_rows.append(row)

        if detect_camera_problem_row(row):
            camera_problem_rows.append(row)

        if detect_risk_row(row):
            risk_rows.append(row)

        missing = detect_empty_fields(row)
        if missing:
            rows_with_missing.append(row)
            for field in missing:
                missing_counter[field] = missing_counter.get(field, 0) + 1

    metrics = []
    add_fact_metric(metrics, "Найдено записей", total, "neutral")

    if overdue_rows:
        add_fact_metric(metrics, "Просрочки", len(overdue_rows), "danger", "Записи с признаками нарушения срока")

    if rows_with_missing:
        add_fact_metric(metrics, "С незаполненными полями", len(rows_with_missing), "warning", "Строки, где есть пустые значимые поля")

    if camera_problem_rows:
        add_fact_metric(metrics, "Проблемы камер", len(camera_problem_rows), "danger", "Нет потока, камера недоступна или не подключена")

    if risk_rows:
        add_fact_metric(metrics, "Риски / нарушения", len(risk_rows), "warning", "Записи с признаками проблемного статуса")

    if len(metrics) == 1:
        add_fact_metric(metrics, "Критичных признаков", 0, "success", "Явных просрочек или пустых полей не найдено")

    top_missing = sorted(missing_counter.items(), key=lambda item: item[1], reverse=True)[:8]

    facts = []

    if overdue_rows:
        facts.append(f"Обнаружено просроченных записей: {len(overdue_rows)}.")

    if rows_with_missing:
        fields_text = ", ".join(f"{name} — {count}" for name, count in top_missing[:5])
        facts.append(f"Есть незаполненные поля: {fields_text}.")

    if camera_problem_rows:
        facts.append(f"Проблемы по камерам/видеопотокам: {len(camera_problem_rows)}.")

    if risk_rows:
        facts.append(f"Записи с признаками риска или нарушения: {len(risk_rows)}.")

    if not facts:
        facts.append("Критичных признаков по найденным записям не выявлено.")

    priority_rows = []

    for collection in [overdue_rows, camera_problem_rows, rows_with_missing, risk_rows, rows]:
        for row in collection:
            if row not in priority_rows:
                priority_rows.append(row)
            if len(priority_rows) >= 6:
                break
        if len(priority_rows) >= 6:
            break

    objects = []

    for row in priority_rows[:6]:
        objects.append({
            "title": object_title_from_row(row),
            "problem": row_problem_text(row),
            "badges": row_badges(row),
        })

    return {
        "module_key": module_info["key"],
        "title": module_info["title"],
        "subtitle": module_info["subtitle"],
        "icon": module_info["icon"],
        "tone": module_info["tone"],
        "source_name": group_name,
        "file": group.get("file", ""),
        "count": total,
        "metrics": metrics,
        "facts": facts,
        "top_missing": [{"field": name, "count": count} for name, count in top_missing],
        "objects": objects,
    }


def prepare_report_for_html(report, selected_blocks=None):
    if not report:
        return report

    if selected_blocks:
        selected_blocks = {b for b in selected_blocks if b in VALID_BLOCK_KEYS}
    if not selected_blocks:
        selected_blocks = None

    prepared = dict(report)
    module_cards = []

    totals = {"records": 0, "overdue": 0, "missing_rows": 0, "camera_problems": 0, "risks": 0}

    for group_name, group in report.get("groups", {}).items():
        module_info = classify_fact_module(group_name, group.get("file", ""))

        if selected_blocks is not None and module_info["key"] not in selected_blocks:
            continue

        card = build_fact_card(module_info, group_name, group)
        module_cards.append(card)
        totals["records"] += card["count"]

        for metric in card["metrics"]:
            label = metric["label"]
            value = int(metric["value"] or 0)
            if label == "Просрочки":
                totals["overdue"] += value
            if label == "С незаполненными полями":
                totals["missing_rows"] += value
            if label == "Проблемы камер":
                totals["camera_problems"] += value
            if label == "Риски / нарушения":
                totals["risks"] += value

    tone_order = {"danger": 0, "warning": 1, "info": 2, "neutral": 3, "success": 4}
    module_cards.sort(key=lambda card: (tone_order.get(card.get("tone"), 9), -int(card.get("count", 0))))

    prepared["module_cards"] = module_cards
    prepared["fact_totals"] = totals
    prepared["blocks"] = REPORT_BLOCKS
    prepared["selected_blocks"] = sorted(selected_blocks) if selected_blocks else []
    return prepared


def should_skip_municipality_report_file(path):
    """
    Исключаем служебные, отладочные и агрегированные файлы.
    Они нужны приложению, но не должны попадать в управленческий отчёт.
    """

    name = path.name.lower()

    if name in EXCLUDED_MUNICIPALITY_REPORT_FILE_NAMES:
        return True

    parts = {part.lower() for part in path.parts}

    if parts & EXCLUDED_MUNICIPALITY_REPORT_PATH_PARTS:
        return True

    rel = str(path).lower()

    bad_fragments = [
        "/debug/",
        "/responses/",
        "/state/",
        "dashboard_state",
        "final_result",
        "raw_response",
        "browser_state",
        "session",
        "cookies",
    ]

    return any(fragment in rel for fragment in bad_fragments)


def normalize_municipality_value(value):
    text = normalize_text(value)

    prefixes = [
        "городской округ",
        "городского округа",
        "го",
        "г о",
        "муниципальный округ",
        "муниципального округа",
        "муниципальное образование",
        "м о",
        "округ",
        "город",
        "г",
    ]

    changed = True

    while changed:
        changed = False

        for prefix in prefixes:
            prefix_norm = normalize_text(prefix)

            if text.startswith(prefix_norm + " "):
                text = text[len(prefix_norm):].strip()
                changed = True

    return text.strip()


def field_name_has_any_hint(field_name, hints):
    key = normalize_text(field_name)

    return any(normalize_text(hint) in key for hint in hints)


def is_municipality_field(field_name):
    return field_name_has_any_hint(field_name, MUNICIPALITY_FIELD_HINTS)


def is_address_field(field_name):
    return field_name_has_any_hint(field_name, ADDRESS_FIELD_HINTS)


def value_looks_like_street_or_address(value, target):
    text = normalize_text(value)
    target_norm = normalize_text(target)

    if not target_norm:
        return False

    address_patterns = [
        f"ул {target_norm}",
        f"улица {target_norm}",
        f"проспект {target_norm}",
        f"пр т {target_norm}",
        f"переулок {target_norm}",
        f"пер {target_norm}",
        f"бульвар {target_norm}",
        f"шоссе {target_norm}",
        f"площадь {target_norm}",
    ]

    if any(pattern in text for pattern in address_patterns):
        return True

    # Например: "г Балашиха, ул Чехова, д. 14"
    if target_norm in text and any(marker in text for marker in [" ул ", " улица ", " д ", " дом ", " корпус "]):
        return True

    return False


def municipality_value_matches(value, municipality_name):
    target = normalize_municipality_value(municipality_name)
    value_norm = normalize_municipality_value(value)

    if not target or not value_norm:
        return False

    if value_looks_like_street_or_address(value, municipality_name):
        return False

    if value_norm == target:
        return True

    # Допускаем форматы типа "городской округ Чехов"
    raw_value = normalize_text(value)

    allowed_phrases = [
        f"городской округ {target}",
        f"го {target}",
        f"г о {target}",
        f"муниципальный округ {target}",
        f"муниципальное образование {target}",
    ]

    if any(phrase == raw_value or phrase in raw_value for phrase in allowed_phrases):
        return True

    return False


def row_matches_municipality(row, municipality_name):
    """
    ВАЖНО:
    Больше не ищем муниципалитет простым substring-поиском по всей строке.
    Иначе 'Чехов' цепляет 'ул. Чехова' в других городах.

    Теперь:
    1. Если есть муниципальные поля — проверяем только их.
    2. Адресные поля не считаем доказательством муниципалитета.
    3. Если муниципальных полей нет — запись не берём, чтобы не ловить улицы.
    """

    target = normalize_text(municipality_name)

    if not target:
        return False

    if not isinstance(row, dict):
        text = safe_str(row)

        if value_looks_like_street_or_address(text, municipality_name):
            return False

        # Для неструктурированных строк оставляем только явный контекст.
        explicit_patterns = [
            f"муниципалитет {target}",
            f"городской округ {target}",
            f"го {target}",
            f"г о {target}",
        ]

        text_norm = normalize_text(text)

        return any(pattern in text_norm for pattern in explicit_patterns)

    municipality_keys = []
    address_keys = []

    for key in row.keys():
        if is_municipality_field(key):
            municipality_keys.append(key)

        if is_address_field(key):
            address_keys.append(key)

    # Если есть отдельные муниципальные поля — доверяем только им.
    if municipality_keys:
        for key in municipality_keys:
            if municipality_value_matches(row.get(key), municipality_name):
                return True

        return False

    # Если муниципальных полей нет, но есть адресные поля,
    # не считаем совпадение по адресу муниципалитетом.
    for key in address_keys:
        value = row.get(key)

        if value_looks_like_street_or_address(value, municipality_name):
            return False

    # Осторожный fallback: ищем только явные подписи "муниципалитет: Чехов"
    # и не принимаем простое наличие слова "Чехов".
    for key, value in row.items():
        key_norm = normalize_text(key)

        if is_address_field(key):
            continue

        if is_technical_fact_key(key):
            continue

        value_norm = normalize_text(value)

        if not value_norm:
            continue

        if value_looks_like_street_or_address(value, municipality_name):
            continue

        explicit_key = any(hint in key_norm for hint in [
            "муниципалитет",
            "городской округ",
            "territory",
            "территория",
        ])

        if explicit_key and municipality_value_matches(value, municipality_name):
            return True

    return False


def value_is_structured_noise(value):
    if isinstance(value, (dict, list, tuple, set)):
        return True

    text = safe_str(value).strip()

    if not text:
        return False

    if text.startswith("{") or text.startswith("["):
        return True

    if '\\"' in text or '"municipality"' in text or "'municipality'" in text:
        return True

    return False


def pick_first_value(row, candidates):
    if not isinstance(row, dict):
        return ""

    normalized_map = {fact_norm(k): k for k in row.keys()}

    for candidate in candidates:
        candidate_norm = fact_norm(candidate)

        for norm_key, original_key in normalized_map.items():
            if candidate_norm == norm_key or candidate_norm in norm_key:
                value = row.get(original_key)

                if value_is_structured_noise(value):
                    continue

                if not is_empty_fact_value(value):
                    return fact_str(value, 180)

    return ""


def object_title_from_row(row):
    title = pick_first_value(row, [
        "organization",
        "организация",
        "org",
        "name",
        "title",
        "object",
        "объект",
        "address",
        "адрес",
        "municipality",
        "муниципалитет",
        "territory",
        "территория",
    ])

    if title:
        return title

    if isinstance(row, dict):
        for key, value in row.items():
            if is_technical_fact_key(key):
                continue

            if value_is_structured_noise(value):
                continue

            if not is_empty_fact_value(value):
                return fact_str(value, 180)

    return "Запись без названия"


def row_problem_text(row):
    if detect_overdue_row(row):
        return "Есть признак просрочки или нарушения срока."

    if detect_camera_problem_row(row):
        return "Есть проблема с камерой или видеопотоком."

    missing = detect_empty_fields(row)
    if missing:
        return "Не заполнены поля: " + ", ".join(missing[:5])

    if detect_risk_row(row):
        return "Есть признаки риска, нарушения или проблемного статуса."

    problem = pick_first_value(row, [
        "problem",
        "issue",
        "reason",
        "comment",
        "проблема",
        "причина",
        "комментарий",
    ])

    if problem:
        return problem

    return "Запись относится к выбранному муниципалитету."


def collect_municipality_data(municipality_name):
    """
    Переопределённый сбор данных:
    - исключает служебные JSON;
    - ищет именно муниципалитет, а не улицу;
    - не цепляет 'ул. Чехова' в других городах.
    """

    groups = {}
    errors = []

    for scan_dir in SCAN_DIRS:
        if not scan_dir.exists():
            continue

        for path in scan_dir.rglob("*"):
            if not path.is_file():
                continue

            if path.suffix.lower() not in ALLOWED_EXTENSIONS:
                continue

            if path.name.startswith("~$"):
                continue

            if should_skip_municipality_report_file(path):
                continue
            # Пропускаем исходный .xlsx, если рядом есть обработанный result.json
            # (одни и те же данные: xlsx — выгрузка из 1С, json — её результат)
            if path.suffix.lower() == ".xlsx" and (path.parent / "result.json").exists():
                continue


            try:
                records = read_records_from_file(path)
            except Exception as e:
                errors.append({
                    "file": str(path.relative_to(BASE_DIR)),
                    "error": str(e),
                })
                continue

            matched = []

            for row in records:
                if row_matches_municipality(row, municipality_name):
                    matched.append(row)

            if matched:
                module_name = detect_module_name(path)
                key = f"{module_name} / {path.name}"

                groups[key] = {
                    "module": module_name,
                    "file": str(path.relative_to(BASE_DIR)),
                    "count": len(matched),
                    "rows": matched,
                }

    total = sum(group["count"] for group in groups.values())

    return {
        "municipality": municipality_name,
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "groups": groups,
        "errors": errors,
        "total": total,
    }


# ============================================================
# MUNICIPALITY ADDRESS REGISTRY
# ============================================================

MUNICIPALITY_REGISTRY_PATH = BASE_DIR / "data" / "municipality_registry.json"
_MUNICIPALITY_REGISTRY_CACHE = None


def load_municipality_registry():
    """
    Загружает справочник муниципалитетов и населённых пунктов.

    Файл:
        data/municipality_registry.json

    Формат:
        {
          "municipalities": {
            "Чехов": {
              "aliases": ["Чехов", "городской округ Чехов"],
              "localities": ["Чехов", "Любучаны", "Столбовая"],
              "exclude_streets": ["улица Чехова", "ул Чехова"]
            }
          }
        }
    """

    global _MUNICIPALITY_REGISTRY_CACHE

    if _MUNICIPALITY_REGISTRY_CACHE is not None:
        return _MUNICIPALITY_REGISTRY_CACHE

    try:
        import json

        if MUNICIPALITY_REGISTRY_PATH.exists():
            _MUNICIPALITY_REGISTRY_CACHE = json.loads(
                MUNICIPALITY_REGISTRY_PATH.read_text(encoding="utf-8")
            )
        else:
            _MUNICIPALITY_REGISTRY_CACHE = {"municipalities": {}}
    except Exception:
        _MUNICIPALITY_REGISTRY_CACHE = {"municipalities": {}}

    return _MUNICIPALITY_REGISTRY_CACHE


def registry_norm(value):
    text = normalize_text(value)
    text = text.replace("ё", "е")
    return text.strip()


def registry_words(value):
    return [part for part in registry_norm(value).split() if part]


@lru_cache(maxsize=1)
def _registry_index():
    """Строит индекс {registry_norm(name): (canonical_name, entry)} один раз."""
    registry = load_municipality_registry()
    municipalities = registry.get("municipalities", {}) or {}
    index = {}
    for canonical_name, entry in municipalities.items():
        names = [canonical_name]
        names.extend(entry.get("aliases", []) or [])
        for name in names:
            key = registry_norm(name)
            if key and key not in index:
                index[key] = (canonical_name, entry)
    return index


def get_registry_municipality_entry(municipality_name):
    target = registry_norm(municipality_name)
    found = _registry_index().get(target)
    if found is not None:
        return found

    # Если точного имени нет в справочнике — возвращаем пустую запись.
    return municipality_name, {
        "aliases": [municipality_name],
        "localities": [municipality_name],
        "exclude_streets": [],
    }


def address_contains_excluded_street(address, entry):
    address_norm = registry_norm(address)

    for street in entry.get("exclude_streets", []) or []:
        street_norm = registry_norm(street)

        if street_norm and street_norm in address_norm:
            return True

    return False


def contains_locality_marker_before(address_norm, locality_norm):
    """
    Проверяет конструкции:
    - г Чехов
    - город Чехов
    - п Любучаны
    - поселок Любучаны
    - д Манушкино
    - деревня Манушкино
    - село Талеж
    - рп Столбовая
    """

    if not locality_norm:
        return False

    markers = [
        "г",
        "город",
        "го",
        "г о",
        "городской округ",
        "п",
        "пос",
        "поселок",
        "посёлок",
        "рп",
        "рабочий поселок",
        "с",
        "село",
        "д",
        "деревня",
        "мкр",
        "микрорайон",
        "тер",
        "территория",
    ]

    for marker in markers:
        pattern = f"{marker} {locality_norm}"

        if pattern in address_norm:
            return True

    return False


def address_starts_with_locality(address_norm, locality_norm):
    """
    Проверяет начало адреса:
    - Чехов, ул. ...
    - Чехов ул ...
    - Любучаны, ...
    """

    if not locality_norm:
        return False

    return (
        address_norm == locality_norm
        or address_norm.startswith(locality_norm + " ")
        or address_norm.startswith(locality_norm + ",")
    )


def address_has_explicit_municipality_alias(address, entry):
    address_norm = registry_norm(address)

    for alias in entry.get("aliases", []) or []:
        alias_norm = registry_norm(alias)

        if not alias_norm:
            continue

        if alias_norm in address_norm:
            return True

    return False


def address_belongs_to_municipality(address, municipality_name):
    """
    Главное правило:
    адрес должен содержать населённый пункт или явное название городского округа
    из справочника.

    При этом "ул. Чехова" не считается населённым пунктом "Чехов".
    """

    if not address:
        return False

    canonical_name, entry = get_registry_municipality_entry(municipality_name)
    address_norm = registry_norm(address)

    if not address_norm:
        return False

    if address_contains_excluded_street(address, entry):
        return False

    # Защита от "ул Чехова" / "улица Чехова" даже без явного exclude.
    target = registry_norm(municipality_name)

    street_false_patterns = [
        f"ул {target}а",
        f"улица {target}а",
        f"пер {target}а",
        f"переулок {target}а",
        f"проспект {target}а",
        f"бульвар {target}а",
        f"шоссе {target}а",
    ]

    if any(pattern in address_norm for pattern in street_false_patterns):
        return False

    # Явные алиасы городского округа.
    # Например: "г.о. Чехов, ..."
    if address_has_explicit_municipality_alias(address, entry):
        return True

    localities = entry.get("localities", []) or []

    for locality in localities:
        locality_norm = registry_norm(locality)

        if not locality_norm:
            continue

        if contains_locality_marker_before(address_norm, locality_norm):
            return True

        if address_starts_with_locality(address_norm, locality_norm):
            return True

    return False


def row_address_belongs_to_municipality(row, municipality_name):
    if not isinstance(row, dict):
        return address_belongs_to_municipality(row, municipality_name)

    address_values = []

    for key, value in row.items():
        if is_technical_fact_key(key):
            continue

        if is_address_field(key):
            address_values.append(value)

    # Если явных адресных полей нет, осторожно пробуем поля с названием/объектом.
    if not address_values:
        for key, value in row.items():
            key_norm = registry_norm(key)

            if any(hint in key_norm for hint in [
                "object",
                "объект",
                "name",
                "название",
                "title",
                "адрес объекта",
                "location",
                "местоположение",
            ]):
                address_values.append(value)

    for value in address_values:
        if value_is_structured_noise(value):
            continue

        if address_belongs_to_municipality(value, municipality_name):
            return True

    return False


def row_matches_municipality(row, municipality_name):
    """
    Финальная логика поиска муниципалитета:

    1. Если есть отдельное поле муниципалитета — проверяем его строго.
    2. Если поля муниципалитета нет — используем адресный справочник.
    3. Простое совпадение слова по всей JSON-строке запрещено.
       Поэтому "ул. Чехова" больше не цепляет Чехов.
    """

    target = registry_norm(municipality_name)

    if not target:
        return False

    if not isinstance(row, dict):
        return address_belongs_to_municipality(row, municipality_name)

    municipality_keys = []

    for key in row.keys():
        if is_municipality_field(key):
            municipality_keys.append(key)

    # Если в данных есть отдельное поле "муниципалитет",
    # оно является главным источником истины.
    if municipality_keys:
        for key in municipality_keys:
            value = row.get(key)

            if municipality_value_matches(value, municipality_name):
                return True

            # Дополнительно разрешаем справочник для случаев:
            # "г.о. Чехов", "городской округ Чехов".
            canonical_name, entry = get_registry_municipality_entry(municipality_name)

            for alias in entry.get("aliases", []) or []:
                if registry_norm(value) == registry_norm(alias):
                    return True

        return False

    # Если поля муниципалитета нет — определяем по адресу через справочник.
    if row_address_belongs_to_municipality(row, municipality_name):
        return True

    return False


# ============================================================
# LIST OF AVAILABLE MUNICIPALITIES (для выпадающего списка)
# ============================================================

def _normalize_display_name(name: str) -> str:
    """Привести 'БОГОРОДСКИЙ' к 'Богородский', а 'дмитров' к 'Дмитров'."""
    s = str(name or "").strip()
    if not s:
        return s
    # Если строка целиком в верхнем регистре кириллицей — переведём в Title-case
    letters = [c for c in s if c.isalpha()]
    if letters and all(c.isupper() for c in letters):
        s = s.capitalize()
    return s


def list_available_municipalities():
    """
    Возвращает отсортированный список муниципалитетов
    строго из справочника data/municipality_registry.json.
    """
    try:
        registry = load_municipality_registry() or {}
    except Exception:
        registry = {}

    names = list((registry.get("municipalities", {}) or {}).keys())

    cleaned = []
    seen = set()
    for raw in names:
        display = _normalize_display_name(raw)
        if not display:
            continue
        # Должна быть хотя бы одна кириллическая буква
        if not re.search(r"[А-Яа-яЁё]", display):
            continue
        # Длина минимум 3 символа
        if len(display) < 3:
            continue
        key = display.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(display)

    cleaned.sort(key=lambda s: s.lower())
    return cleaned

# ============================================================
